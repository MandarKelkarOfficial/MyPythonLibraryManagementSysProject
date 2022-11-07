# Original Program by Mandar Kelkar
from datetime import *
from tkinter import *
from tkinter import messagebox
from tkinter import filedialog
from openpyxl import *
import mysql.connector


# Creating a funcion to get user data through database and display it
def UserdataDisplay():
    
    global myresult
    search_id = reg_no_entry.get()
    mysqlDb = mysql.connector.connect(host="localhost",user="goditachi",password="maddy06",database="saneguruji_data")
    mycursor=mysqlDb.cursor()
    try:
        mycursor.execute("SELECT * FROM sanegurugidata where RegNo  =  '" + search_id + "'")
        myresult = mycursor.fetchall()
 
        for x in myresult:
            print(x)
        name_entry.delete(0, END)
        name_entry.insert(END, x[0])
        
        address_entry.delete(0, END)
        address_entry.insert(END, x[1])
        
        PhoneNo_entry.delete(0, END)
        PhoneNo_entry.insert(END, x[2])
        
        previous_credit_entry.delete(0, END)
        previous_credit_entry.insert(END, x[3])
        
        april_entry.delete(0, END)
        april_entry.insert(END, x[4])
        
        may_entry.delete(0, END)
        may_entry.insert(END, x[5])
        
        june_entry.delete(0, END)
        june_entry.insert(END, x[6])
        
        july_entry.delete(0, END)
        july_entry.insert(END, x[7])
        
        august_entry.delete(0, END)
        august_entry.insert(END, x[8])
        
        september_entry.delete(0, END)
        september_entry.insert(END, x[9])
        
        october_entry.delete(0, END)
        october_entry.insert(END, x[10])
        
        diwali_anka_entry.delete(0, END)
        diwali_anka_entry.insert(END, x[11])
        
        nov_entry.delete(0, END)
        nov_entry.insert(END, x[12])
        
        dec_entry.delete(0, END)
        dec_entry.insert(END, x[13])
        
        jan_entry.delete(0, END)
        jan_entry.insert(END, x[14])
        
        feb_entry.delete(0, END)
        feb_entry.insert(END, x[15])
        
        march_entry.delete(0, END)
        march_entry.insert(END, x[16])
        
        # reg_no_entry.delete(0, END)
        # reg_no_entry.insert(END, x[17])
        
        book_name_entry.delete(0, END)
        book_name_entry.insert(END, x[18])
        
        bookauthor_name_entry.delete(0, END)
        bookauthor_name_entry.insert(END, x[19])
        
        signature_id_entry.delete(0, END)
        signature_id_entry.insert(END, x[20])
        
        date_entry.delete(0, END)
        date_entry.insert(END, x[21])
        
        librarian_sign_entry.delete(0, END)
        librarian_sign_entry.insert(END, x[22])
        
        book_rtn_date_entry.delete(0, END)
        book_rtn_date_entry.insert(END, x[23])
        
        remark_entry.delete(0, END)
        remark_entry.insert(END, x[24])
 
    except Exception as e:
       print(e)
       mysqlDb.rollback()
       mysqlDb.close()
    

# Globally declared the path of excel sheet
sg_data = load_workbook("C:\\Users\\manda\\OneDrive\\Desktop\\python_app\\SaneGurugi_software\\SG_ClientData.xlsx")
sheet = sg_data.active

#sg_client_data = load_workbook("C:\\Users\\manda\\OneDrive\\Desktop\\python_app\\SaneGurugi_software\\SG_ClientData.xlsx")
#sheet1 = sg_client_data.active

# Function to load first data sheet
def GetTable():
    	
	# resize the width of columns in
	# excel spreadsheet
    sheet.column_dimensions['A'].width = 40
    sheet.column_dimensions['B'].width = 50
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 50
    sheet.column_dimensions['H'].width = 50
    sheet.column_dimensions['I'].width = 50
    sheet.column_dimensions['J'].width = 50
    sheet.column_dimensions['K'].width = 50
    sheet.column_dimensions['L'].width = 50
    sheet.column_dimensions['M'].width = 50
    sheet.column_dimensions['N'].width = 50
    sheet.column_dimensions['O'].width = 10
    sheet.column_dimensions['P'].width = 50
    sheet.column_dimensions['Q'].width = 50
    sheet.column_dimensions['R'].width = 50
    sheet.column_dimensions['S'].width = 30
    sheet.column_dimensions['T'].width = 50
    sheet.column_dimensions['U'].width = 50
    sheet.column_dimensions['V'].width = 50
    sheet.column_dimensions['W'].width = 50
    sheet.column_dimensions['X'].width = 50
    sheet.column_dimensions['Y'].width = 50


	# write given data to an excel spreadsheet
	# at particular location
    sheet.cell(row=1, column=1).value = "Name "
    sheet.cell(row=1, column=2).value = "Address "
    sheet.cell(row=1, column=3).value = "Contact No "
    sheet.cell(row=1, column=4).value = "Pending Amount "
    sheet.cell(row=1, column=5).value = "April"
    sheet.cell(row=1, column=6).value = "May"
    sheet.cell(row=1, column=7).value = "June"
    sheet.cell(row=1, column=8).value = "July"
    sheet.cell(row=1, column=9).value = "August"
    sheet.cell(row=1, column=10).value = "September"
    sheet.cell(row=1, column=11).value = "October"
    sheet.cell(row=1, column=12).value = "Diwali Anka"
    sheet.cell(row=1, column=13).value = "November"
    sheet.cell(row=1, column=14).value = "December"
    sheet.cell(row=1, column=15).value = "Janury"
    sheet.cell(row=1, column=16).value = "Februry"
    sheet.cell(row=1, column=17).value = "March"
    sheet.cell(row=1, column=18).value = "Reg no "
    sheet.cell(row=1, column=19).value = "Book name "
    sheet.cell(row=1, column=20).value = "Book author "
    sheet.cell(row=1, column=21).value = "Signature id"
    sheet.cell(row=1, column=22).value = "Book Fetching Date"
    sheet.cell(row=1, column=23).value = "Librarian Sign"
    sheet.cell(row=1, column=24).value = "Book Returning Date"
    sheet.cell(row=1, column=25).value = "Remark"


# function to clear the data in entry
def clearAll():
    
    # clear the content of text entry box
    previous_credit_entry.delete(0, END)
    name_entry.delete(0, END)
    address_entry.delete(0, END)
    PhoneNo_entry.delete(0, END)
    april_entry.delete(0, END)
    may_entry.delete(0, END)
    june_entry.delete(0, END)
    july_entry.delete(0, END)
    august_entry.delete(0, END)
    september_entry.delete(0, END)
    october_entry.delete(0, END)
    diwali_anka_entry.delete(0, END)
    nov_entry.delete(0, END)
    dec_entry.delete(0, END)
    jan_entry.delete(0, END)
    feb_entry.delete(0, END)
    march_entry.delete(0, END)
    reg_no_entry.delete(0, END)
    book_name_entry.delete(0, END)
    bookauthor_name_entry.delete(0, END)
    signature_id_entry.delete(0, END)
    date_entry.delete(0, END)
    librarian_sign_entry.delete(0, END)
    book_rtn_date_entry.delete(0, END)
    remark_entry.delete(0, END)


 
 
 
# Function to insert data in sheet
def input():
    # If all input fields are empty then show empty input
    if (name_entry.get() == "" and
       address_entry.get == "" and
       previous_credit_entry.get()== "" and
       april_entry.get() == "" and
       may_entry.get() == "" and
       june_entry.get() == "" and
       july_entry.get()== "" and
       august_entry.get()=="" and
       diwali_anka_entry.get()==""and
       nov_entry.get()=="" and 
       dec_entry.get()=="" and
       jan_entry.get()=="" and
       feb_entry.get()=="" and
       march_entry.get=="" and
       reg_no_entry.get() == "" and
       book_name_entry.get()=="" and
       bookauthor_name_entry.get()== "" and
       signature_id_entry.get()==""):
        
         messagebox.showinfo("Error", "Form is Empty")
    else:
        
        # assigning the max row and max column
		# value upto which data is written
		# in an excel sheet to the variable
        current_row = sheet.max_row
        current_column = sheet.max_column
        
        # get method returns current text
		# as string which we write into
		# excel spreadsheet at particular location
        sheet.cell(row=current_row + 1, column=1).value = name_entry.get()
        sheet.cell(row=current_row + 1, column=2).value = address_entry.get()
        sheet.cell(row=current_row + 1, column=3).value = PhoneNo_entry.get()
        sheet.cell(row=current_row + 1, column=4).value = previous_credit_entry.get()
        sheet.cell(row=current_row + 1, column=5).value = april_entry.get()
        sheet.cell(row=current_row + 1, column=6).value = may_entry.get()
        sheet.cell(row=current_row + 1, column=7).value = june_entry.get()
        sheet.cell(row=current_row + 1, column=8).value = july_entry.get()
        sheet.cell(row=current_row + 1, column=9).value = august_entry.get()
        sheet.cell(row=current_row + 1, column=10).value = september_entry.get()
        sheet.cell(row=current_row + 1, column=11).value = october_entry.get()
        sheet.cell(row=current_row + 1, column=12).value = diwali_anka_entry.get()
        sheet.cell(row=current_row + 1, column=13).value = nov_entry.get()
        sheet.cell(row=current_row + 1, column=14).value = dec_entry.get()
        sheet.cell(row=current_row + 1, column=15).value = jan_entry.get()
        sheet.cell(row=current_row + 1, column=16).value = feb_entry.get()
        sheet.cell(row=current_row + 1, column=17).value = march_entry.get()
        sheet.cell(row=current_row + 1, column=18).value = reg_no_entry.get()
        sheet.cell(row=current_row + 1, column=19).value = book_name_entry.get()
        sheet.cell(row=current_row + 1, column=20).value = bookauthor_name_entry.get()
        sheet.cell(row=current_row + 1, column=21).value = signature_id_entry.get()
        sheet.cell(row=current_row + 1, column=22).value = date_entry.get()
        sheet.cell(row=current_row + 1, column=23).value = librarian_sign_entry.get()
        sheet.cell(row=current_row + 1, column=24).value = book_rtn_date_entry.get()
        sheet.cell(row=current_row + 1, column=25).value = remark_entry.get()
        

		# save the file
        sg_data.save("C:\\Users\\manda\\OneDrive\\Desktop\\python_app\\SaneGurugi_software\\SG_ClientData.xlsx")
        
        # call clear function to clear the table after saving
        clearAll()

        
"""def insert_today_date():
    raw_TS = datetime.now(IST)
    formatted_now = raw_TS.strftime ("%d-%m-%Y")
    date_entry.set (formatted_now)"""
    
    
"""
def GetTable1():
    sheet1.column_dimensions['A'].width = 10
    sheet1.column_dimensions['B'].width = 50
    sheet1.column_dimensions['C'].width = 50
    sheet1.column_dimensions['D'].width = 50
    sheet1.column_dimensions['E'].width = 30
    sheet1.column_dimensions['F'].width = 50
    
    sheet1.cell(row=1, column=1).value = "Reg no "
    sheet1.cell(row=1, column=2).value = "Book name "
    sheet1.cell(row=1, column=3).value = "Book author "
    sheet1.cell(row=1, column=4).value = "Signature id"
    sheet1.cell(row=1, column=5).value = "Date"
    sheet1.cell(row=1, column=6).value = "Remark"



def input1():
    if( reg_no_entry.get() == "" and
       book_name_entry.get()=="" and
       bookauthor_name_entry.get()== "" and
       signature_id_entry.get()==""):
        
        messagebox.showinfo("Empty boxes")
    
    else:
       
       current_row1 = sheet.max_row
       current_column1 = sheet.max_column
       
       sheet1.cell(row=current_row1 + 1, column=1).value = reg_no_entry.get()
       sheet1.cell(row=current_row1 + 1, column=2).value = book_name_entry.get()
       sheet1.cell(row=current_row1 + 1, column=3).value = bookauthor_name_entry.get()
       sheet1.cell(row=current_row1 + 1, column=4).value = signature_id_entry.get()
       sheet1.cell(row=current_row1 + 1, column=5).value = date_entry.get()
       sheet1.cell(row=current_row1 + 1, column=6).value = remark_entry.get()
       
       sg_client_data.save("C:\\Users\\manda\\OneDrive\\Desktop\\python_app\\SaneGurugi_software\\SG_ClientData.xlsx")
       clearAll()"""

       
        
def newWindow():
    
    def clearWindow():
        setpath_entry.delete(0, END)
    
    def setWindowPath():
        path = filedialog.askdirectory()
        setpath_label.config(text=path)
        
        
    nw = Tk()
    nw.title("Set Path")
    nw.config(borderwidth=10, bg="#8A2BE2")
    nw.geometry("350x500")
    nw.minsize(350, 500)
    nw.maxsize(350, 500)
    
    
    # Label and Entry for setting path of xl file
    
    setpath_label = Label(nw, text="Set path of excel file", bg="#8A2BE2", fg="white", font="Arial 15 bold")
    setpath_label.place(x="60", y="100")
    setpath_entry = Entry(nw, borderwidth=0, bg="white", fg="#9400D3", font="helvetica 10 italic", width="35")
    setpath_entry.place(x="45", y="150")
    
    # Button for executing command to set xl file
    
    setpath_button = Button(nw, text="SET", borderwidth=0, bg="#9400D3", fg="white", font="Arial 9 bold", width=15, command=setWindowPath)
    setpath_button.place(x="40", y="190")
    setpathClear_button = Button(nw, text="CLEAR", borderwidth=0, bg="#9400D3", fg="white", font="Arial 9 bold", width=15, command=clearWindow)
    setpathClear_button.place(x="180", y="190")
    
    
    
    nw.mainloop()
    


def printFile():
    messagebox.showinfo("","Print")

    
#=======================================================================================================================================================================================================

root = Tk()   # Creating main frame
root.title("SaneGurugiDoc")
root.geometry("1000x700")
root.minsize(1000, 700)
root.maxsize(1000, 700)
root.config(bg="#8A2BE2", borderwidth=10) # Setting color of main frame


# Heading labels

heading_label1 = Label(root, font="Arial 11 bold", text="Baristor. Nath pai Sevangan organized ",bg="#8A2BE2", fg="black")
heading_label1.place(x="385", y="10")
heading_label2 = Label(root,text="Sane Guruji Vachan Mandir, Malvan",font="Arial 15 bold",bg="#8A2BE2", fg="black")
heading_label2.place(x="350", y="35")

#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# Label and Entry for participants adress and name

name_label  = Label(root, text="Participants name :",bg="#8A2BE2", fg="white", font="Arial 10 bold")
name_label.place(x="40", y="90")
name_entry = Entry(root, borderwidth=0, bg="white",fg="#9400D3", width="50")
name_entry.place(x="180", y="90")

address_label  = Label(root, text="Address                 :",bg="#8A2BE2", fg="white", font="Arial 10 bold")
address_label.place(x="40", y="120")
address_entry = Entry(root, borderwidth=0, bg="white",fg="#9400D3", width="50")
address_entry.place(x="180", y="120")

PhoneNo_label  = Label(root, text="Contact No            :",bg="#8A2BE2", fg="white", font="Arial 10 bold")
PhoneNo_label.place(x="40", y="150")
PhoneNo_entry = Entry(root, borderwidth=0, bg="white",fg="#9400D3", width="50")
PhoneNo_entry.place(x="180", y="150")



#------------------------------------------------------------------------------------------------------------------------

# some labels and entrys more participnts' data

previous_credit_label = Label(root,text="Pending(₹) :", bg="#8A2BE2", fg="white", font="Arial 10 bold")
previous_credit_label.place(x="40", y="195")
previous_credit_entry = Entry(root,borderwidth=0, bg="white",fg="#9400D3",font="Arial 8 bold", width="12")
previous_credit_entry.place(x=40, y="220")

april_label = Label(root, text="April", bg="#8A2BE2", fg="white", font="Arial 10 bold")
april_label.place(x="160", y="195")
april_entry = Entry(root,borderwidth=0, bg="white",fg="#9400D3", font="Arial 8 bold",width="12")
april_entry.place(x="160", y="220")

may_label = Label(root, text="May ", bg="#8A2BE2", fg="white", font="Arial 10 bold")
may_label.place(x="280", y="195")
may_entry = Entry(root,borderwidth=0, bg="white",fg="#9400D3",font="Arial 8 bold", width="12")
may_entry.place(x="280", y="220")

june_label = Label(root, text="June", bg="#8A2BE2", fg="white", font="Arial 10 bold")
june_label.place(x="400", y="195")
june_entry = Entry(root,borderwidth=0, bg="white",fg="#9400D3",font="Arial 8 bold", width="12")
june_entry.place(x="400", y="220")

july_label = Label(root, text="July", bg="#8A2BE2", fg="white", font="Arial 10 bold")
july_label.place(x="520", y="195")
july_entry = Entry(root,borderwidth=0, bg="white",fg="#9400D3",font="Arial 8 bold", width="12")
july_entry.place(x="520", y="220")

august_label = Label(root, text="august", bg="#8A2BE2", fg="white", font="Arial 10 bold")
august_label.place(x="640", y="195")
august_entry = Entry(root,borderwidth=0, bg="white",fg="#9400D3",font="Arial 8 bold", width="12")
august_entry.place(x="640", y="220")

september_label = Label(root, text="sept", bg="#8A2BE2", fg="white", font="Arial 10 bold")
september_label.place(x="760", y="195")
september_entry = Entry(root,borderwidth=0, bg="white",fg="#9400D3",font="Arial 8 bold", width="12")
september_entry.place(x="760", y="220")

# Above 7
#_________________________________________________________________________________________________________________________________________________________________
# Below &

october_label = Label(root, text="Oct", bg="#8A2BE2", fg="white", font="Arial 10 bold")
october_label.place(x="40", y="265")
october_entry = Entry(root,borderwidth=0, bg="white",fg="#9400D3",font="Arial 8 bold", width="12")
october_entry.place(x="40", y="290")

diwali_anka_label = Label(root, text="Diwali Anka", bg="#8A2BE2", fg="white", font="Arial 10 bold")
diwali_anka_label.place(x="160", y="265")
diwali_anka_entry = Entry(root,borderwidth=0, bg="white",fg="#9400D3",font="Arial 8 bold", width="12")
diwali_anka_entry.place(x="160", y="290")

nov_label = Label(root, text="Nov", bg="#8A2BE2", fg="white", font="Arial 10 bold")
nov_label.place(x="280", y="265")
nov_entry = Entry(root,borderwidth=0, bg="white",fg="#9400D3",font="Arial 8 bold", width="12")
nov_entry.place(x="280", y="290")

dec_label = Label(root, text="Dec", bg="#8A2BE2", fg="white", font="Arial 10 bold")
dec_label.place(x="400", y="265")
dec_entry = Entry(root,borderwidth=0, bg="white",fg="#9400D3",font="Arial 8 bold", width="12")
dec_entry.place(x="400", y="290")

jan_label = Label(root, text="Jan", bg="#8A2BE2", fg="white", font="Arial 10 bold")
jan_label.place(x="520", y="265")
jan_entry = Entry(root,borderwidth=0, bg="white",fg="#9400D3",font="Arial 8 bold", width="12")
jan_entry.place(x="520", y="290")

feb_label = Label(root, text="Feb", bg="#8A2BE2", fg="white", font="Arial 10 bold")
feb_label.place(x="640", y="265")
feb_entry = Entry(root,borderwidth=0, bg="white",fg="#9400D3",font="Arial 8 bold", width="12")
feb_entry.place(x="640", y="290")

march_label = Label(root, text="March", bg="#8A2BE2", fg="white", font="Arial 10 bold")
march_label.place(x="760", y="265")
march_entry = Entry(root,borderwidth=0, bg="white",fg="#9400D3",font="Arial 8 bold", width="12")
march_entry.place(x="760", y="290")

# _______________________________________________________________________________________________________________________

# A table for some data
reg_no_label = Label(root, text="Reg no" , bg="#8A2BE2", fg="white", font="Arial 10 bold")
reg_no_label.place(x="40", y="360")
reg_no_entry = Entry(root,borderwidth=0, bg="white",fg="#9400D3",font="Arial 10 bold", width="10")
reg_no_entry.place(x="150", y="360")

book_name_label = Label(root, text="Book name" , bg="#8A2BE2", fg="white", font="Arial 10 bold")
book_name_label.place(x="40", y="410")
book_name_entry = Entry(root,borderwidth=0, bg="white",fg="#9400D3",font="Arial 10 bold", width="20")
book_name_entry.place(x="150", y="410")

bookauthor_name_label = Label(root, text="Book Author " , bg="#8A2BE2", fg="white", font="Arial 10 bold")
bookauthor_name_label.place(x="40", y="460")
bookauthor_name_entry = Entry(root,borderwidth=0, bg="white",fg="#9400D3",font="Arial 10 bold", width="20")
bookauthor_name_entry.place(x="150", y="460")

signature_id_label = Label(root, text=" Signature" , bg="#8A2BE2", fg="white", font="Arial 10 bold")
signature_id_label.place(x="40", y="510")
signature_id_entry = Entry(root,borderwidth=0, bg="white",fg="#9400D3",font="Arial 10 bold", width="10")
signature_id_entry.place(x="150", y="510")

date_label = Label(root, text="Book Fetching Date " , bg="#8A2BE2", fg="white", font="Arial 10 bold")
date_label.place(x="400", y="410")
date_entry = Entry(root,borderwidth=0, bg="white",fg="#9400D3",font="Arial 10 bold", width="10")
date_entry.place(x="550", y="410")

librarian_sign_label = Label(root, text="Librarian Signature " , bg="#8A2BE2", fg="white", font="Arial 10 bold")
librarian_sign_label.place(x="400", y="460")
librarian_sign_entry = Entry(root,borderwidth=0, bg="white",fg="#9400D3",font="Arial 10 bold", width="10")
librarian_sign_entry.place(x="550", y="460")

book_rtn_date_label = Label(root, text="Book Returning Date " , bg="#8A2BE2", fg="white", font="Arial 10 bold")
book_rtn_date_label.place(x="400", y="510")
book_rtn_date_entry = Entry(root,borderwidth=0, bg="white",fg="#9400D3",font="Arial 10 bold", width="10")
book_rtn_date_entry.place(x="550", y="510")

remark_label = Label(root, text="Remark " , bg="#8A2BE2", fg="white", font="Arial 10 bold")
remark_label.place(x="680", y="410")
remark_entry = Entry(root,borderwidth=0, bg="white",fg="#9400D3",font="Arial 10 bold", width="15")
remark_entry.place(x="750", y="410")


# A submit button to submit the data and a clear button to clear a data

submit_data = Button(root, text="SUBMIT", font="Arial 12 bold", width=25, borderwidth=0, bg="#4B0082", fg="white", command=input)
submit_data.place(x="40", y="580")
clear_data = Button(root, text="CLEAR", font="Arial 12 bold", width=25, borderwidth=0, bg="#4B0082", fg="white", command=clearAll)
clear_data.place(x="300", y="580")
search_data = Button(root, text="SEARCH", font="Arial 12 bold", width=25, borderwidth=0, bg="#4B0082", fg="white", command=UserdataDisplay)
search_data.place(x="560", y="580")
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Adding more details to root

 

#-------------------------------------------------------------------------------------------------------------------------------------------------------
# Creating a menu
# root window
# create a menubar
menubar = Menu(root, bg="#4B0082",fg="white")
root.config(menu=menubar)


# create the file_menu
file_menu = Menu(menubar,tearoff=0, bg="#4B0082",fg="white")

# add menu items to the File menu
file_menu.add_command(label='New', command=newWindow)
file_menu.add_command(label='Open')
file_menu.add_command(label='Close')
file_menu.add_command(label="PRINT", command=printFile)
file_menu.add_separator()

# add a submenu
sub_menu = Menu(file_menu, tearoff=0, bg="#4B0082",fg="white")
sub_menu.add_command(label='Keyboard Shortcuts')
sub_menu.add_command(label='Color Themes')

# add the File menu to the menubar
file_menu.add_cascade(label="Preferences",menu=sub_menu)


# add Exit menu item
file_menu.add_separator()
file_menu.add_command(label='Exit',command=root.destroy)


menubar.add_cascade(label="File",menu=file_menu,underline=0)
# create the Help menu
help_menu = Menu(menubar,tearoff=0, bg="#4B0082",fg="white")

help_menu.add_command(label='Welcome')
help_menu.add_command(label='About...')

# add the Help menu to the menubar
menubar.add_cascade(label="Help",menu=help_menu, underline=0)

    
# Calling excel function written on line no 8 and 55
GetTable()


root.mainloop()